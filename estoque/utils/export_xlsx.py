"""
Utilitário para exportação de dados para planilhas Excel (XLSX)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from decimal import Decimal


def exportar_produtos_para_xlsx(produtos, nome_arquivo='produtos.xlsx'):
    """
    Exporta lista de produtos para arquivo XLSX.
    
    Args:
        produtos: QuerySet ou lista de produtos
        nome_arquivo: Nome do arquivo a ser gerado
        
    Returns:
        HttpResponse com o arquivo XLSX
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"
    
    # Cabeçalho
    headers = ['Código', 'Nome', 'Categoria', 'Unidade', 'Quantidade', 'Custo Unitário', 'Valor Total']
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados
    for row, produto in enumerate(produtos, 2):
        ws.cell(row=row, column=1).value = produto.codigo
        ws.cell(row=row, column=2).value = produto.nome
        ws.cell(row=row, column=3).value = produto.categoria.nome if produto.categoria else ''
        ws.cell(row=row, column=4).value = produto.get_unidade_display()
        ws.cell(row=row, column=5).value = float(produto.quantidade_estoque)
        ws.cell(row=row, column=6).value = float(produto.custo_unitario)
        ws.cell(row=row, column=7).value = float(produto.valor_total_estoque)
        
        # Formatação numérica
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        ws.cell(row=row, column=6).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=7).number_format = 'R$ #,##0.00'
    
    # Ajusta largura das colunas
    column_widths = [15, 40, 20, 12, 12, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    
    wb.save(response)
    return response


def exportar_relatorio_para_xlsx(movimentacoes, resumo, nome_arquivo='relatorio_estoque.xlsx'):
    """
    Exporta relatório de movimentações para arquivo XLSX.
    
    Args:
        movimentacoes: QuerySet ou lista de movimentações
        resumo: Dicionário com resumo (entradas, saidas, saldo_final)
        nome_arquivo: Nome do arquivo a ser gerado
        
    Returns:
        HttpResponse com o arquivo XLSX
    """
    wb = Workbook()
    
    # Planilha de Movimentações
    ws = wb.active
    ws.title = "Movimentações"
    
    headers = ['Data', 'Tipo', 'Produto', 'Quantidade', 'Unidade', 'Custo Unitário', 'Usuário']
    
    # Cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados
    for row, mov in enumerate(movimentacoes, 2):
        ws.cell(row=row, column=1).value = mov.created_at.strftime('%d/%m/%Y %H:%M')
        ws.cell(row=row, column=2).value = mov.get_tipo_display()
        ws.cell(row=row, column=3).value = mov.produto.nome
        ws.cell(row=row, column=4).value = float(mov.quantidade)
        ws.cell(row=row, column=5).value = mov.produto.get_unidade_display()
        ws.cell(row=row, column=6).value = float(mov.custo_unitario) if mov.custo_unitario else 0
        ws.cell(row=row, column=7).value = mov.usuario.username if mov.usuario else ''
        
        # Formatação
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=6).number_format = 'R$ #,##0.00'
    
    # Ajusta largura das colunas
    column_widths = [18, 10, 40, 12, 10, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Planilha de Resumo
    ws_resumo = wb.create_sheet("Resumo")
    
    resumo_headers = ['Item', 'Valor']
    for col, header in enumerate(resumo_headers, 1):
        cell = ws_resumo.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws_resumo.cell(row=2, column=1).value = 'Total de Entradas'
    ws_resumo.cell(row=2, column=2).value = float(resumo.get('entradas', 0))
    ws_resumo.cell(row=2, column=2).number_format = '#,##0.00'
    
    ws_resumo.cell(row=3, column=1).value = 'Total de Saídas'
    ws_resumo.cell(row=3, column=2).value = float(resumo.get('saidas', 0))
    ws_resumo.cell(row=3, column=2).number_format = '#,##0.00'
    
    ws_resumo.cell(row=4, column=1).value = 'Saldo Final'
    ws_resumo.cell(row=4, column=2).value = float(resumo.get('saldo_final', 0))
    ws_resumo.cell(row=4, column=2).number_format = '#,##0.00'
    
    # Ajusta largura das colunas do resumo
    ws_resumo.column_dimensions['A'].width = 20
    ws_resumo.column_dimensions['B'].width = 15
    
    # Resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    
    wb.save(response)
    return response

