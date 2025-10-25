from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from django.utils.dateparse import parse_date
import io

from .models import Product
from apps.stock.models import StockMovement
from .serializers import ProductSerializer


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_products_excel(request):
    """Exporta lista de produtos para Excel"""
    
    # Buscar todos os produtos
    products = Product.objects.select_related('category').all()
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"
    
    # Cabeçalhos
    headers = [
        'ID', 'Nome', 'Descrição', 'SKU', 'Categoria', 
        'Preço', 'Quantidade', 'Quantidade Mínima', 'Status'
    ]
    
    # Estilizar cabeçalhos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dados dos produtos
    for row, product in enumerate(products, 2):
        status = "Estoque Baixo" if product.quantity <= product.min_quantity else "Normal"
        
        ws.cell(row=row, column=1, value=product.id)
        ws.cell(row=row, column=2, value=product.name)
        ws.cell(row=row, column=3, value=product.description)
        ws.cell(row=row, column=4, value=product.sku)
        ws.cell(row=row, column=5, value=product.category.name)
        ws.cell(row=row, column=6, value=float(product.price))
        ws.cell(row=row, column=7, value=product.quantity)
        ws.cell(row=row, column=8, value=product.min_quantity)
        ws.cell(row=row, column=9, value=status)
    
    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
    
    # Criar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="produtos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Salvar workbook na resposta
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_movements_excel(request):
    """Exporta movimentações de estoque para Excel com filtros de data"""
    
    # Obter filtros de data dos parâmetros
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Buscar movimentações com filtros
    movements = StockMovement.objects.select_related('product', 'user').all().order_by('-created_at')
    
    # Aplicar filtros de data se fornecidos
    if start_date:
        start_date_obj = parse_date(start_date)
        if start_date_obj:
            movements = movements.filter(created_at__date__gte=start_date_obj)
    
    if end_date:
        end_date_obj = parse_date(end_date)
        if end_date_obj:
            movements = movements.filter(created_at__date__lte=end_date_obj)
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    
    # Definir título baseado nos filtros
    if start_date and end_date:
        ws.title = f"Movimentações {start_date} a {end_date}"
    elif start_date:
        ws.title = f"Movimentações desde {start_date}"
    elif end_date:
        ws.title = f"Movimentações até {end_date}"
    else:
        ws.title = "Movimentações"
    
    # Cabeçalhos
    headers = [
        'ID', 'Produto', 'SKU', 'Categoria', 'Tipo', 'Quantidade', 
        'Motivo', 'Observações', 'Usuário', 'Data'
    ]
    
    # Estilizar cabeçalhos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dados das movimentações
    for row, movement in enumerate(movements, 2):
        tipo_map = {
            'in': 'Entrada',
            'out': 'Saída',
            'adjustment': 'Ajuste'
        }
        
        ws.cell(row=row, column=1, value=movement.id)
        ws.cell(row=row, column=2, value=movement.product.name)
        ws.cell(row=row, column=3, value=movement.product.sku)
        ws.cell(row=row, column=4, value=movement.product.category.name)
        ws.cell(row=row, column=5, value=tipo_map.get(movement.movement_type, movement.movement_type))
        ws.cell(row=row, column=6, value=movement.quantity)
        ws.cell(row=row, column=7, value=movement.reason)
        ws.cell(row=row, column=8, value=movement.notes)
        ws.cell(row=row, column=9, value=movement.user.full_name)
        ws.cell(row=row, column=10, value=movement.created_at.strftime("%d/%m/%Y %H:%M"))
    
    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
    
    # Criar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nome do arquivo baseado nos filtros
    if start_date and end_date:
        filename = f'movimentacoes_{start_date}_a_{end_date}.xlsx'
    elif start_date:
        filename = f'movimentacoes_desde_{start_date}.xlsx'
    elif end_date:
        filename = f'movimentacoes_ate_{end_date}.xlsx'
    else:
        filename = f'movimentacoes_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Salvar workbook na resposta
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_inventory_report_excel(request):
    """Exporta relatório completo de inventário para Excel"""
    
    # Buscar produtos e movimentações
    products = Product.objects.select_related('category').all()
    movements = StockMovement.objects.select_related('product', 'user').all().order_by('-created_at')
    
    # Criar workbook com múltiplas abas
    wb = Workbook()
    
    # Aba 1: Resumo de Produtos
    ws_products = wb.active
    ws_products.title = "Resumo Produtos"
    
    headers_products = [
        'ID', 'Nome', 'SKU', 'Categoria', 'Preço', 
        'Quantidade Atual', 'Quantidade Mínima', 'Status', 'Valor Total'
    ]
    
    # Estilizar cabeçalhos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers_products, 1):
        cell = ws_products.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dados dos produtos
    total_value = 0
    for row, product in enumerate(products, 2):
        status = "Estoque Baixo" if product.quantity <= product.min_quantity else "Normal"
        valor_total = float(product.price) * product.quantity
        total_value += valor_total
        
        ws_products.cell(row=row, column=1, value=product.id)
        ws_products.cell(row=row, column=2, value=product.name)
        ws_products.cell(row=row, column=3, value=product.sku)
        ws_products.cell(row=row, column=4, value=product.category.name)
        ws_products.cell(row=row, column=5, value=float(product.price))
        ws_products.cell(row=row, column=6, value=product.quantity)
        ws_products.cell(row=row, column=7, value=product.min_quantity)
        ws_products.cell(row=row, column=8, value=status)
        ws_products.cell(row=row, column=9, value=valor_total)
    
    # Adicionar linha de total
    total_row = len(products) + 3
    ws_products.cell(row=total_row, column=8, value="VALOR TOTAL DO ESTOQUE:")
    ws_products.cell(row=total_row, column=9, value=total_value)
    
    # Ajustar largura das colunas
    for col in range(1, len(headers_products) + 1):
        column_letter = get_column_letter(col)
        ws_products.column_dimensions[column_letter].width = 20
    
    # Aba 2: Movimentações Recentes
    ws_movements = wb.create_sheet("Movimentações")
    
    headers_movements = [
        'Data', 'Produto', 'SKU', 'Tipo', 'Quantidade', 
        'Motivo', 'Usuário'
    ]
    
    for col, header in enumerate(headers_movements, 1):
        cell = ws_movements.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dados das movimentações (últimas 50)
    recent_movements = movements[:50]
    for row, movement in enumerate(recent_movements, 2):
        tipo_map = {
            'in': 'Entrada',
            'out': 'Saída',
            'adjustment': 'Ajuste'
        }
        
        ws_movements.cell(row=row, column=1, value=movement.created_at.strftime("%d/%m/%Y %H:%M"))
        ws_movements.cell(row=row, column=2, value=movement.product.name)
        ws_movements.cell(row=row, column=3, value=movement.product.sku)
        ws_movements.cell(row=row, column=4, value=tipo_map.get(movement.movement_type, movement.movement_type))
        ws_movements.cell(row=row, column=5, value=movement.quantity)
        ws_movements.cell(row=row, column=6, value=movement.reason)
        ws_movements.cell(row=row, column=7, value=movement.user.full_name)
    
    # Ajustar largura das colunas
    for col in range(1, len(headers_movements) + 1):
        column_letter = get_column_letter(col)
        ws_movements.column_dimensions[column_letter].width = 20
    
    # Criar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="relatorio_inventario_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Salvar workbook na resposta
    wb.save(response)
    return response
